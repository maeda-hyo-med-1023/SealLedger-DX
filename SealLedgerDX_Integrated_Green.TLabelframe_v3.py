import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font
import os
import shutil
import re
import fnmatch
from PIL import Image
import math
import sys
from copy import copy

# ------------------------------------------------------------
# OCR用（pytesseract）
# ------------------------------------------------------------
try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
except ImportError:
    PYTESSERACT_AVAILABLE = False
    print("警告: pytesseractがインストールされていません。")
    print("OCR機能を使用するには: pip install pytesseract")
    print("また、Tesseract-OCR本体のインストールが必要です。")

# ------------------------------------------------------------
# 共通定数 / ヘルパー関数
# ------------------------------------------------------------
MM_PER_INCH = 25.4

def mm_to_px(mm, dpi):
    return int(mm / MM_PER_INCH * dpi)

def pt_to_px(pt, dpi):
    return int(pt / 72 * dpi)

def char_to_px(char_count, dpi=96):
    base_px_per_char = 7 * (dpi / 96)
    return int(char_count * base_px_per_char)

# Excel操作用（openpyxl）
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("警告: openpyxlがインストールされていません。")
    print("画像挿入機能を使用するには: pip install openpyxl")

# PDF変換用（pdf2image）
try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False
    print("警告: pdf2imageがインストールされていません。")
    print("PDF変換機能を使用するには: pip install pdf2image")
    print("また、popplerのインストールが必要です。")

# ------------------------------------------------------------
# セルスタイルコピー用ヘルパー関数
# ------------------------------------------------------------
def copy_cell_style(source_cell, target_cell):
    """ソースセルのスタイル（罫線・フォント・塗りつぶし・配置など）をターゲットセルにコピー"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

# ------------------------------------------------------------
# メインアプリケーション
# ------------------------------------------------------------
class SealLedgerDXApp:
    def __init__(self, root):
        self.root = root
        root.title("SealLedger DX - 統合自動化システム v3")
        root.geometry("980x800")
        root.minsize(900, 720)
        root.resizable(True, True)
        root.configure(bg="#E8F5E9")

        # ---- ★★★ 動的フォント設定（メイリオ） ★★★ ----
        self.base_font = font.Font(family="Meiryo", size=10)
        self.bold_font = font.Font(family="Meiryo", size=10, weight="bold")
        self.console_font = font.Font(family="Consolas", size=9)
        self.default_font_tuple = ("Meiryo", 10)
        root.option_add("*Font", self.base_font)

        # ---- ★★★ フォントサイズコントロール（大・中・小） ★★★ ----
        self.create_font_controls()

        # ---- ttkスタイル設定（緑色枠＋各タブの背景色スタイル） ----
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("Green.TLabelframe", borderwidth=3, relief="solid", bordercolor="#2E7D32")
        self.style.configure("Green.TLabelframe.Label", font=self.bold_font)
        self.style.configure("Green.TButton", background="#4CAF50", foreground="white", borderwidth=1, focusthickness=3, focuscolor="none")
        self.style.map("Green.TButton", background=[("active", "#66BB6A")])

        # ---- 各タブの背景色に対応するLabelFrameスタイルを作成 ----
        tab_colors = [
            "#E8F5E9",  # ① 薄緑
            "#FFF3E0",  # ② 薄橙
            "#E3F2FD",  # ③ 薄青
            "#FCE4EC",  # ④ 薄桃
            "#F3E5F5"   # ⑤ 薄紫
        ]
        self.labelframe_styles = []
        for i, color in enumerate(tab_colors):
            style_name = f"Green{i}.TLabelframe"
            self.style.configure(style_name, background=color, borderwidth=3, relief="solid", bordercolor="#2E7D32")
            self.style.configure(f"{style_name}.Label", font=self.bold_font)
            self.labelframe_styles.append(style_name)

        # ---- シルバー光彩 ----
        self.create_silver_border()

        # ---- ノートブック ----
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=20, pady=20)

        # --- タブ1（tk.Frame） ---
        self.tab1 = tk.Frame(self.notebook, bg=tab_colors[0])
        self.notebook.add(self.tab1, text="① フォーマット＆入力規則")
        self.setup_tab1(tab_colors[0], self.labelframe_styles[0])

        # --- タブ2 ---
        self.tab2 = tk.Frame(self.notebook, bg=tab_colors[1])
        self.notebook.add(self.tab2, text="② PDF変換＆フォルダ作成")
        self.setup_tab2(tab_colors[1], self.labelframe_styles[1])

        # --- タブ3 ---
        self.tab3 = tk.Frame(self.notebook, bg=tab_colors[2])
        self.notebook.add(self.tab3, text="③ ファイル振り分け")
        self.setup_tab3(tab_colors[2], self.labelframe_styles[2])

        # --- タブ4 ---
        self.tab4 = tk.Frame(self.notebook, bg=tab_colors[3])
        self.notebook.add(self.tab4, text="④ 校印台帳カット支援")
        self.setup_tab4(tab_colors[3], self.labelframe_styles[3])

        # --- タブ5 ---
        self.tab5 = tk.Frame(self.notebook, bg=tab_colors[4])
        self.notebook.add(self.tab5, text="⑤ OCR（画像→テキスト化）")
        self.setup_tab5(tab_colors[4], self.labelframe_styles[4])

        # ステータスバー
        self.status_var = tk.StringVar(value="準備完了")
        self.status_bar = ttk.Label(root, textvariable=self.status_var, relief="sunken", anchor="w", font=self.base_font)
        self.status_bar.pack(side="bottom", fill="x")

    # ============================================================
    # ★★★ フォントサイズコントロール（大・中・小） ★★★
    # ============================================================
    def create_font_controls(self):
        """フォントサイズ切替パネルを生成（メインノートブックの上部に配置）"""
        control_frame = tk.Frame(self.root, bg="#E8F5E9", padx=10, pady=5)
        control_frame.pack(side="top", fill="x", padx=10)

        tk.Label(control_frame, text="表示サイズ：", bg="#E8F5E9", font=self.base_font).pack(side="left", padx=(0, 10))

        self.font_size_var = tk.StringVar(value="中")
        sizes = [("小", 8), ("中", 10), ("大", 14)]

        for label, size in sizes:
            rb = tk.Radiobutton(
                control_frame,
                text=label,
                variable=self.font_size_var,
                value=label,
                bg="#E8F5E9",
                font=self.base_font,
                command=lambda s=size: self.set_font_size(s)
            )
            rb.pack(side="left", padx=5)

        # 現在のサイズを表示するラベル（フィードバック用）
        self.size_indicator = tk.Label(control_frame, text="(現在: 10pt)", bg="#E8F5E9", fg="#555555", font=self.base_font)
        self.size_indicator.pack(side="left", padx=15)

    def set_font_size(self, size):
        """アプリケーション全体のフォントサイズを変更"""
        self.base_font.config(size=size)
        self.bold_font.config(size=size)
        console_size = max(8, size - 1)
        self.console_font.config(size=console_size)

        self.status_bar.config(font=self.base_font)
        self.size_indicator.config(text=f"(現在: {size}pt)")

        self.style.configure("TNotebook.Tab", font=self.base_font)
        self.style.configure("Green.TButton", font=self.base_font)
        self.style.configure("Green.TLabelframe.Label", font=self.bold_font)
        for style_name in self.labelframe_styles:
            self.style.configure(f"{style_name}.Label", font=self.bold_font)

        self.status_var.set(f"フォントサイズを {size}pt に変更しました")

    # ============================================================
    # シルバー光彩
    # ============================================================
    def create_silver_border(self):
        """ウィンドウ内側にシルバー光彩の装飾ラインを描画（キャンバス）"""
        self.border_canvas = tk.Canvas(self.root, bg="#E8F5E9", highlightthickness=0, height=6)
        self.border_canvas.pack(side="top", fill="x", padx=0, pady=(0,0))
        for x in range(0, 1000, 2):
            brightness = 200 - abs(x - 500) * 0.2
            if brightness < 150:
                brightness = 150
            color = f"#{int(brightness):02x}{int(brightness):02x}{int(brightness):02x}"
            self.border_canvas.create_line(x, 2, x+2, 4, fill=color, width=2)
        self.border_canvas.create_line(0, 4, 1000, 4, fill="#C0C0C0", width=1)
        self.border_canvas.create_line(0, 1, 1000, 1, fill="#FFFFFF", width=2)

        self.border_canvas_bottom = tk.Canvas(self.root, bg="#E8F5E9", highlightthickness=0, height=6)
        self.border_canvas_bottom.pack(side="bottom", fill="x", padx=0, pady=(0,0))
        for x in range(0, 1000, 2):
            brightness = 200 - abs(x - 500) * 0.2
            if brightness < 150:
                brightness = 150
            color = f"#{int(brightness):02x}{int(brightness):02x}{int(brightness):02x}"
            self.border_canvas_bottom.create_line(x, 2, x+2, 4, fill=color, width=2)
        self.border_canvas_bottom.create_line(0, 4, 1000, 4, fill="#C0C0C0", width=1)
        self.border_canvas_bottom.create_line(0, 1, 1000, 1, fill="#FFFFFF", width=2)

    # ============================================================
    # タブ1: フォーマット＆入力規則
    # ============================================================
    def setup_tab1(self, bg_color, label_style):
        parent = self.tab1
        parent.configure(bg=bg_color)
        pad = 10

        # ...（①-1 の部分は変更なし）...

        frame2 = ttk.LabelFrame(parent, text="①-2 入力規則を自動セット", style=label_style, padding=pad)
        frame2.pack(fill="x", padx=pad, pady=pad)

        row2 = tk.Frame(frame2, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="マスターExcel：").pack(side="left")
        self.val_master = tk.StringVar()
        ttk.Entry(row2, textvariable=self.val_master, width=30, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row2, text="参照", command=lambda: self.browse_file(self.val_master, "Excel files", ".xlsx")).pack(side="left")

        row3 = tk.Frame(frame2, bg=bg_color)
        row3.pack(fill="x", pady=2)
        ttk.Label(row3, text="マスターシート名：").pack(side="left")
        self.val_master_sheet = tk.StringVar(value="マスタリスト")
        ttk.Entry(row3, textvariable=self.val_master_sheet, width=15, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row3, text="マスター列：").pack(side="left", padx=10)
        self.val_master_col = tk.StringVar(value="A")
        ttk.Entry(row3, textvariable=self.val_master_col, width=5, font=self.base_font).pack(side="left")

        row4 = tk.Frame(frame2, bg=bg_color)
        row4.pack(fill="x", pady=2)
        ttk.Label(row4, text="対象Excel：").pack(side="left")
        self.val_target = tk.StringVar()
        ttk.Entry(row4, textvariable=self.val_target, width=30, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row4, text="参照", command=lambda: self.browse_file(self.val_target, "Excel files", ".xlsx")).pack(side="left")

        row5 = tk.Frame(frame2, bg=bg_color)
        row5.pack(fill="x", pady=2)
        ttk.Label(row5, text="対象シート名：").pack(side="left")
        self.val_target_sheet = tk.StringVar(value="校印使用台帳")
        ttk.Entry(row5, textvariable=self.val_target_sheet, width=15, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row5, text="設定列：").pack(side="left", padx=10)
        self.val_target_col = tk.StringVar(value="E")
        ttk.Entry(row5, textvariable=self.val_target_col, width=5, font=self.base_font).pack(side="left")
        ttk.Label(row5, text="開始行：").pack(side="left", padx=10)
        self.val_start_row = tk.IntVar(value=3)
        ttk.Entry(row5, textvariable=self.val_start_row, width=5, font=self.base_font).pack(side="left")
        ttk.Label(row5, text="終了行（0=最終行まで）：").pack(side="left", padx=10)
        self.val_end_row = tk.IntVar(value=0)
        ttk.Entry(row5, textvariable=self.val_end_row, width=5, font=self.base_font).pack(side="left")
        ttk.Button(row5, text="▶ 入力規則セット", command=self.run_validation_set, style="Green.TButton").pack(side="left", padx=10)

    # ★★★ 修正点：同じシート内にフォーマット罫線を連続追加 ★★★
    def run_format_copy(self):
        template = self.fmt_template.get().strip()
        if not template or not os.path.exists(template):
            messagebox.showerror("エラー", "テンプレートファイルを選択してください。")
            return
        pages = self.fmt_pages.get()
        if pages < 1:
            messagebox.showerror("エラー", "ページ数は1以上を指定してください。")
            return

        dir_name = os.path.dirname(template)
        base_name = os.path.splitext(os.path.basename(template))[0]
        output = os.path.join(dir_name, f"{base_name}_generated.xlsx")
        if os.path.exists(output):
            if not messagebox.askyesno("確認", f"出力先\n{output}\nは既に存在します。上書きしますか？"):
                return

        try:
            shutil.copy2(template, output)
            wb = openpyxl.load_workbook(output)

            target_sheet = None
            for sheet_name in wb.sheetnames:
                if '校印使用台帳' in sheet_name:
                    target_sheet = wb[sheet_name]
                    break
            if target_sheet is None:
                target_sheet = wb[wb.sheetnames[0]]

            # テンプレートブロックの行範囲
            template_start_row = 1
            template_end_row = 23
            block_height = template_end_row - template_start_row + 1

            # 列範囲を動的に取得（最大列数まで）
            max_col = target_sheet.max_column

            current_max_row = target_sheet.max_row
            if current_max_row < template_end_row:
                current_max_row = template_end_row

            insert_start_row = current_max_row + 1

            from openpyxl.utils import get_column_letter

            for page_idx in range(pages):
                dest_start_row = insert_start_row + (page_idx * block_height)

                # セル値＋スタイルをコピー（全列）
                for row_offset in range(block_height):
                    src_row = template_start_row + row_offset
                    dest_row = dest_start_row + row_offset
                    for col in range(1, max_col + 1):
                        src_cell = target_sheet.cell(row=src_row, column=col)
                        dest_cell = target_sheet.cell(row=dest_row, column=col)
                        dest_cell.value = src_cell.value
                        copy_cell_style(src_cell, dest_cell)

                # 行の高さをコピー
                for row_offset in range(block_height):
                    src_row = template_start_row + row_offset
                    dest_row = dest_start_row + row_offset
                    if target_sheet.row_dimensions[src_row].height:
                        target_sheet.row_dimensions[dest_row].height = target_sheet.row_dimensions[src_row].height

                # セルの結合をコピー（全列対応）
                for merged_range in list(target_sheet.merged_cells.ranges):
                    if merged_range.min_row <= template_end_row and merged_range.max_row <= template_end_row:
                        row_offset = merged_range.min_row - template_start_row
                        new_min_row = dest_start_row + row_offset
                        new_max_row2 = dest_start_row + (merged_range.max_row - template_start_row)
                        new_min_col = merged_range.min_col
                        new_max_col = merged_range.max_col
                        target_sheet.merge_cells(
                            start_row=new_min_row,
                            start_column=new_min_col,
                            end_row=new_max_row2,
                            end_column=new_max_col
                        )

            # 印刷範囲を全列に拡張
            new_max_row = insert_start_row + (pages * block_height) - 1
            last_col_letter = get_column_letter(max_col)
            target_sheet.print_area = f"A1:{last_col_letter}{new_max_row}"

            wb.save(output)
            self.status_var.set(f"✅ 複製完了: {pages}ページを同じシート内に追加しました。")
            messagebox.showinfo("成功", f"完了！\n{output}\n同じシート内に{pages}ページ分のフォーマットを追加しました。")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    def run_validation_set(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("エラー", "openpyxlがインストールされていません。")
            return
        master_file = self.val_master.get().strip()
        target_file = self.val_target.get().strip()
        if not master_file or not target_file:
            messagebox.showerror("エラー", "マスターと対象のファイルを指定してください。")
            return
        try:
            wb_master = openpyxl.load_workbook(master_file, read_only=True)
            master_sheet_name = self.val_master_sheet.get().strip()
            if master_sheet_name not in wb_master.sheetnames:
                messagebox.showerror("エラー", f"マスターシート '{master_sheet_name}' が見つかりません。")
                return
            ws_master = wb_master[master_sheet_name]
            master_col = self.val_master_col.get().strip().upper()
            col_idx = openpyxl.utils.column_index_from_string(master_col)
            values = []
            for row in ws_master.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                if row[0] is not None and str(row[0]).strip() != "":
                    values.append(str(row[0]).strip())
            wb_master.close()
            if not values:
                messagebox.showerror("エラー", "マスター列に値がありません。")
                return

            wb_target = openpyxl.load_workbook(target_file)
            target_sheet_name = self.val_target_sheet.get().strip()
            if target_sheet_name not in wb_target.sheetnames:
                messagebox.showerror("エラー", f"対象シート '{target_sheet_name}' が見つかりません。")
                return
            ws_target = wb_target[target_sheet_name]
            target_col = self.val_target_col.get().strip().upper()
            start_row = self.val_start_row.get()
            end_row = ws_target.max_row
            if end_row < start_row:
                end_row = start_row

            from openpyxl.worksheet.datavalidation import DataValidation
            dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
            dv.error = 'リストから選択してください。'
            dv.errorTitle = '入力エラー'
            ws_target.add_data_validation(dv)
            dv.add(f"{target_col}{start_row}:{target_col}{end_row}")

            wb_target.save(target_file)
            self.status_var.set(f"✅ 入力規則セット完了: {len(values)}件")
            messagebox.showinfo("成功", f"{len(values)}件の選択肢を設定しました。")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    # ============================================================
    # タブ2: PDF変換＆フォルダ作成
    # ============================================================
    def setup_tab2(self, bg_color, label_style):
        parent = self.tab2
        parent.configure(bg=bg_color)
        pad = 10

        frame1 = ttk.LabelFrame(parent, text="②-1 PDF → 画像変換", style=label_style, padding=pad)
        frame1.pack(fill="x", padx=pad, pady=pad)

        row1 = tk.Frame(frame1, bg=bg_color)
        row1.pack(fill="x", pady=2)
        ttk.Label(row1, text="PDFファイル：").pack(side="left")
        self.pdf_path = tk.StringVar()
        ttk.Entry(row1, textvariable=self.pdf_path, width=35, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row1, text="参照", command=lambda: self.browse_file(self.pdf_path, "PDF files", ".pdf")).pack(side="left")
        ttk.Label(row1, text="出力先：").pack(side="left", padx=10)
        self.pdf_output = tk.StringVar()
        ttk.Entry(row1, textvariable=self.pdf_output, width=25, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row1, text="参照", command=lambda: self.browse_folder(self.pdf_output)).pack(side="left")
        row2 = tk.Frame(frame1, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="DPI：").pack(side="left")
        self.pdf_dpi = tk.IntVar(value=300)
        ttk.Entry(row2, textvariable=self.pdf_dpi, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row2, text="ページ範囲（例: 1-10, 空欄=全ページ）：").pack(side="left", padx=10)
        self.pdf_range = tk.StringVar()
        ttk.Entry(row2, textvariable=self.pdf_range, width=15, font=self.base_font).pack(side="left")
        ttk.Button(row2, text="▶ 変換実行", command=self.run_pdf_convert, style="Green.TButton").pack(side="left", padx=10)

        frame2 = ttk.LabelFrame(parent, text="②-2 フォルダ一括作成", style=label_style, padding=pad)
        frame2.pack(fill="x", padx=pad, pady=pad)

        row3 = tk.Frame(frame2, bg=bg_color)
        row3.pack(fill="x", pady=2)
        ttk.Label(row3, text="ベース名：").pack(side="left")
        self.folder_base = tk.StringVar(value="page_")
        ttk.Entry(row3, textvariable=self.folder_base, width=15, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row3, text="作成数：").pack(side="left", padx=10)
        self.folder_count = tk.IntVar(value=62)
        ttk.Spinbox(row3, from_=1, to=999, textvariable=self.folder_count, width=8).pack(side="left")
        ttk.Label(row3, text="出力先：").pack(side="left", padx=10)
        self.folder_output = tk.StringVar()
        ttk.Entry(row3, textvariable=self.folder_output, width=25, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row3, text="参照", command=lambda: self.browse_folder(self.folder_output)).pack(side="left")
        ttk.Button(row3, text="▶ フォルダ作成", command=self.run_folder_create, style="Green.TButton").pack(side="left", padx=10)

    def run_pdf_convert(self):
        if not PDF2IMAGE_AVAILABLE:
            messagebox.showerror("エラー", "pdf2imageがインストールされていません。\npip install pdf2image")
            return
        pdf = self.pdf_path.get().strip()
        out = self.pdf_output.get().strip()
        if not pdf or not out:
            messagebox.showerror("エラー", "PDFと出力先を指定してください。")
            return
        if not os.path.exists(out):
            os.makedirs(out, exist_ok=True)
        dpi = self.pdf_dpi.get()
        try:
            self.status_var.set("PDF変換中...")
            images = convert_from_path(pdf, dpi=dpi)
            total = len(images)
            for i, img in enumerate(images, 1):
                img.save(os.path.join(out, f"page_{i}.png"), "PNG")
                self.status_var.set(f"変換中: {i}/{total}")
                self.root.update()
            self.status_var.set(f"✅ PDF変換完了: {total}ページ")
            messagebox.showinfo("成功", f"{total}ページを変換しました。")
        except Exception as e:
            messagebox.showerror("エラー", f"変換エラー: {e}")

    def run_folder_create(self):
        base = self.folder_base.get().strip()
        count = self.folder_count.get()
        out = self.folder_output.get().strip()
        if not base or not out:
            messagebox.showerror("エラー", "ベース名と出力先を指定してください。")
            return
        try:
            for i in range(1, count + 1):
                os.makedirs(os.path.join(out, f"{base}{i}"), exist_ok=True)
            self.status_var.set(f"✅ フォルダ作成完了: {count}個")
            messagebox.showinfo("成功", f"{count}個のフォルダを作成しました。")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    # ============================================================
    # タブ3: ファイル振り分け
    # ============================================================
    def setup_tab3(self, bg_color, label_style):
        parent = self.tab3
        parent.configure(bg=bg_color)
        pad = 10
        frame = ttk.LabelFrame(parent, text="③ ファイル→同名フォルダ自動振り分け", style=label_style, padding=pad)
        frame.pack(fill="both", expand=True, padx=pad, pady=pad)

        row1 = tk.Frame(frame, bg=bg_color)
        row1.pack(fill="x", pady=2)
        ttk.Label(row1, text="ベース名：").pack(side="left")
        self.router_base = tk.StringVar(value="page_")
        ttk.Entry(row1, textvariable=self.router_base, width=15, font=self.base_font).pack(side="left", padx=5)

        row2 = tk.Frame(frame, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="ソースフォルダ（ファイル）：").pack(side="left")
        self.router_source = tk.StringVar()
        ttk.Entry(row2, textvariable=self.router_source, width=35, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row2, text="参照", command=lambda: self.browse_folder(self.router_source)).pack(side="left")

        row3 = tk.Frame(frame, bg=bg_color)
        row3.pack(fill="x", pady=2)
        ttk.Label(row3, text="ターゲットフォルダ（格納先）：").pack(side="left")
        self.router_target = tk.StringVar()
        ttk.Entry(row3, textvariable=self.router_target, width=35, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row3, text="参照", command=lambda: self.browse_folder(self.router_target)).pack(side="left")

        row4 = tk.Frame(frame, bg=bg_color)
        row4.pack(fill="x", pady=2)
        ttk.Label(row4, text="末尾付与名称（例：_processed）：").pack(side="left")
        self.router_suffix = tk.StringVar(value="")
        ttk.Entry(row4, textvariable=self.router_suffix, width=15, font=self.base_font).pack(side="left", padx=5)
        self.router_auto_create = tk.BooleanVar(value=True)
        ttk.Checkbutton(row4, text="フォルダ自動作成", variable=self.router_auto_create).pack(side="left", padx=10)
        ttk.Button(row4, text="▶ 一括振り分け", command=self.run_router, style="Green.TButton").pack(side="left", padx=10)

        self.router_log = tk.Text(frame, height=8, font=self.console_font, state="disabled", bg="#fafafa")
        self.router_log.pack(fill="both", expand=True, padx=5, pady=5)

    def router_log_msg(self, msg):
        self.router_log.config(state="normal")
        self.router_log.insert(tk.END, msg + "\n")
        self.router_log.see(tk.END)
        self.router_log.config(state="disabled")
        self.root.update()

    def run_router(self):
        source = self.router_source.get().strip()
        target = self.router_target.get().strip()
        base = self.router_base.get().strip()
        suffix = self.router_suffix.get().strip()
        auto_create = self.router_auto_create.get()

        if not source or not target:
            messagebox.showerror("エラー", "ソースとターゲットフォルダを指定してください。")
            return
        if not os.path.exists(source):
            messagebox.showerror("エラー", "ソースフォルダが存在しません。")
            return
        self.router_log.config(state="normal")
        self.router_log.delete(1.0, tk.END)
        self.router_log.config(state="disabled")

        try:
            files = [f for f in os.listdir(source) if os.path.isfile(os.path.join(source, f))]
            target_files = [f for f in files if f.startswith(base)]
            if not target_files:
                messagebox.showinfo("情報", f"「{base}」で始まるファイルが見つかりません。")
                return

            copied = 0
            for filename in target_files:
                folder_name = os.path.splitext(filename)[0]
                src_path = os.path.join(source, filename)
                dest_folder = os.path.join(target, folder_name)

                if not os.path.exists(dest_folder):
                    if auto_create:
                        os.makedirs(dest_folder, exist_ok=True)
                        self.router_log_msg(f"📁 フォルダ作成: {folder_name}")
                    else:
                        self.router_log_msg(f"⚠️ スキップ: {filename} (フォルダなし)")
                        continue

                name, ext = os.path.splitext(filename)
                if suffix:
                    new_name = f"{name}{suffix}{ext}"
                else:
                    new_name = filename

                counter = 1
                dest_path = os.path.join(dest_folder, new_name)
                while os.path.exists(dest_path):
                    name_base, ext2 = os.path.splitext(new_name)
                    dest_path = os.path.join(dest_folder, f"{name_base} ({counter}){ext2}")
                    counter += 1

                shutil.copy2(src_path, dest_path)
                copied += 1
                self.router_log_msg(f"✅ {filename} → {folder_name}/{os.path.basename(dest_path)}")

            self.status_var.set(f"✅ 振り分け完了: {copied}件")
            messagebox.showinfo("成功", f"{copied}件のファイルを振り分けました。")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    # ============================================================
    # タブ4: 校印台帳カット支援
    # ============================================================
    def setup_tab4(self, bg_color, label_style):
        parent = self.tab4
        parent.configure(bg=bg_color)
        pad = 10

        common_frame = ttk.LabelFrame(parent, text="共通設定", style=label_style, padding=pad)
        common_frame.pack(fill="x", padx=pad, pady=pad)

        row1 = tk.Frame(common_frame, bg=bg_color)
        row1.pack(fill="x", pady=2)
        ttk.Label(row1, text="Excelファイル：").pack(side="left")
        self.t4_excel = tk.StringVar()
        ttk.Entry(row1, textvariable=self.t4_excel, width=40, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row1, text="参照", command=lambda: self.browse_file(self.t4_excel, "Excel files", ".xlsx")).pack(side="left")

        row2 = tk.Frame(common_frame, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="画像フォルダ：").pack(side="left")
        self.t4_image_folder = tk.StringVar()
        ttk.Entry(row2, textvariable=self.t4_image_folder, width=40, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row2, text="参照", command=lambda: self.browse_folder(self.t4_image_folder)).pack(side="left")
        ttk.Button(row2, text="📁 _cut作成", command=lambda: self.make_folder(self.t4_image_folder.get(), "_cut")).pack(side="left", padx=5)
        ttk.Button(row2, text="📁 resize作成", command=lambda: self.make_folder(self.t4_image_folder.get(), "resize")).pack(side="left", padx=5)

        row3 = tk.Frame(common_frame, bg=bg_color)
        row3.pack(fill="x", pady=2)
        ttk.Label(row3, text="ページ番号：").pack(side="left")
        self.t4_page = tk.IntVar(value=33)
        ttk.Spinbox(row3, from_=1, to=999, textvariable=self.t4_page, width=8).pack(side="left", padx=5)
        ttk.Button(row3, text="▶ 次ページへ (+1)", command=self.t4_update_page, style="Green.TButton").pack(side="left", padx=10)

        sub_notebook = ttk.Notebook(parent)
        sub_notebook.pack(fill="both", expand=True, padx=pad, pady=pad)

        self.sub1 = tk.Frame(sub_notebook, bg=bg_color)
        sub_notebook.add(self.sub1, text="① 使途・文書番号")
        self.setup_t4_sub1(bg_color)

        self.sub2 = tk.Frame(sub_notebook, bg=bg_color)
        sub_notebook.add(self.sub2, text="② 日付・使用印 (D列)")
        self.setup_t4_sub2(bg_color)

        self.sub3 = tk.Frame(sub_notebook, bg=bg_color)
        sub_notebook.add(self.sub3, text="③ 使用部署・取扱者 (H列)")
        self.setup_t4_sub3(bg_color)

        log_frame = ttk.LabelFrame(parent, text="処理ログ", padding=pad)
        log_frame.pack(fill="both", expand=True, padx=pad, pady=pad)
        self.t4_log = tk.Text(log_frame, height=6, font=self.console_font, state="disabled", bg="#fafafa")
        self.t4_log.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.t4_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.t4_log.config(yscrollcommand=scrollbar.set)

    def t4_log_msg(self, msg):
        self.t4_log.config(state="normal")
        self.t4_log.insert(tk.END, msg + "\n")
        self.t4_log.see(tk.END)
        self.t4_log.config(state="disabled")
        self.root.update()

    def t4_update_page(self):
        current = self.t4_page.get()
        self.t4_page.set(current + 1)
        self.t4_log_msg(f"📄 ページ番号を {current} → {current + 1} に更新")

    def make_folder(self, base_path, folder_name):
        if not base_path:
            messagebox.showerror("エラー", "画像フォルダを先に指定してください。")
            return
        path = os.path.join(base_path, folder_name)
        os.makedirs(path, exist_ok=True)
        self.t4_log_msg(f"✅ フォルダ作成: {path}")
        messagebox.showinfo("成功", f"作成しました: {path}")

    def setup_t4_sub1(self, bg_color):
        parent = self.sub1
        parent.configure(bg=bg_color)
        pad = 10
        frame = ttk.LabelFrame(parent, text="カット設定 (cm)", padding=pad)
        frame.pack(fill="x", padx=pad, pady=pad)
        row = tk.Frame(frame, bg=bg_color)
        row.pack(fill="x", pady=2)
        ttk.Label(row, text="上：").pack(side="left", padx=(0,5))
        self.t4_cut1_top = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut1_top, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="下：").pack(side="left", padx=(10,5))
        self.t4_cut1_bottom = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut1_bottom, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="左：").pack(side="left", padx=(10,5))
        self.t4_cut1_left = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut1_left, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="右：").pack(side="left", padx=(10,5))
        self.t4_cut1_right = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut1_right, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row, text="✂ カット実行 (使途・文書番号)", command=self.t4_exec_cut_purpose, style="Green.TButton").pack(side="left", padx=10)

    def t4_exec_cut_purpose(self):
        self._t4_generic_cut("_cutPurposeDocumentNo.png", self.t4_cut1_top, self.t4_cut1_bottom, self.t4_cut1_left, self.t4_cut1_right, resize=False)

    def setup_t4_sub2(self, bg_color):
        parent = self.sub2
        parent.configure(bg=bg_color)
        pad = 10
        cut_frame = ttk.LabelFrame(parent, text="カット設定 (cm)", padding=pad)
        cut_frame.pack(fill="x", padx=pad, pady=pad)
        row = tk.Frame(cut_frame, bg=bg_color)
        row.pack(fill="x", pady=2)
        ttk.Label(row, text="上：").pack(side="left", padx=(0,5))
        self.t4_cut2_top = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut2_top, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="下：").pack(side="left", padx=(10,5))
        self.t4_cut2_bottom = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut2_bottom, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="左：").pack(side="left", padx=(10,5))
        self.t4_cut2_left = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut2_left, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="右：").pack(side="left", padx=(10,5))
        self.t4_cut2_right = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut2_right, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row, text="✂ カット実行 (日付・使用印)", command=self.t4_exec_cut_datestamp, style="Green.TButton").pack(side="left", padx=10)

        resize_frame = ttk.LabelFrame(parent, text="リサイズ設定 (ピクセル)", padding=pad)
        resize_frame.pack(fill="x", padx=pad, pady=pad)
        row2 = tk.Frame(resize_frame, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="幅：").pack(side="left", padx=(0,5))
        self.t4_resize_w2 = tk.IntVar(value=200)
        ttk.Entry(row2, textvariable=self.t4_resize_w2, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row2, text="高さ：").pack(side="left", padx=(10,5))
        self.t4_resize_h2 = tk.IntVar(value=200)
        ttk.Entry(row2, textvariable=self.t4_resize_h2, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row2, text="🔄 リサイズ実行", command=self.t4_exec_resize_datestamp, style="Green.TButton").pack(side="left", padx=10)
        ttk.Button(row2, text="📌 D列に挿入", command=lambda: self.t4_insert_image("D", "datestamp"), style="Green.TButton").pack(side="left", padx=10)

    def t4_exec_cut_datestamp(self):
        self._t4_generic_cut("_cutDate_Stamp.png", self.t4_cut2_top, self.t4_cut2_bottom, self.t4_cut2_left, self.t4_cut2_right, resize=False)

    def t4_exec_resize_datestamp(self):
        self._t4_generic_resize("_cutDate_Stamp.png", "_resizeDateStamp.png", self.t4_resize_w2, self.t4_resize_h2)

    def setup_t4_sub3(self, bg_color):
        parent = self.sub3
        parent.configure(bg=bg_color)
        pad = 10
        cut_frame = ttk.LabelFrame(parent, text="カット設定 (cm)", padding=pad)
        cut_frame.pack(fill="x", padx=pad, pady=pad)
        row = tk.Frame(cut_frame, bg=bg_color)
        row.pack(fill="x", pady=2)
        ttk.Label(row, text="上：").pack(side="left", padx=(0,5))
        self.t4_cut3_top = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut3_top, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="下：").pack(side="left", padx=(10,5))
        self.t4_cut3_bottom = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut3_bottom, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="左：").pack(side="left", padx=(10,5))
        self.t4_cut3_left = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut3_left, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row, text="右：").pack(side="left", padx=(10,5))
        self.t4_cut3_right = tk.DoubleVar(value=0.0)
        ttk.Entry(row, textvariable=self.t4_cut3_right, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row, text="✂ カット実行 (使用部署・取扱者)", command=self.t4_exec_cut_department, style="Green.TButton").pack(side="left", padx=10)

        resize_frame = ttk.LabelFrame(parent, text="リサイズ設定 (ピクセル)", padding=pad)
        resize_frame.pack(fill="x", padx=pad, pady=pad)
        row2 = tk.Frame(resize_frame, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="幅：").pack(side="left", padx=(0,5))
        self.t4_resize_w3 = tk.IntVar(value=200)
        ttk.Entry(row2, textvariable=self.t4_resize_w3, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row2, text="高さ：").pack(side="left", padx=(10,5))
        self.t4_resize_h3 = tk.IntVar(value=200)
        ttk.Entry(row2, textvariable=self.t4_resize_h3, width=6, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row2, text="🔄 リサイズ実行", command=self.t4_exec_resize_department, style="Green.TButton").pack(side="left", padx=10)
        ttk.Button(row2, text="📌 H列に挿入", command=lambda: self.t4_insert_image("H", "department"), style="Green.TButton").pack(side="left", padx=10)

    def t4_exec_cut_department(self):
        self._t4_generic_cut("_cutDepartment_User.png", self.t4_cut3_top, self.t4_cut3_bottom, self.t4_cut3_left, self.t4_cut3_right, resize=False)

    def t4_exec_resize_department(self):
        self._t4_generic_resize("_cutDepartment_User.png", "_resizeDepartmentUser.png", self.t4_resize_w3, self.t4_resize_h3)

    def _t4_generic_cut(self, suffix_filename, top_var, bottom_var, left_var, right_var, resize=False):
        folder = self.t4_image_folder.get().strip()
        if not folder:
            messagebox.showerror("エラー", "画像フォルダを選択してください。")
            return
        cut_folder = os.path.join(folder, "_cut")
        os.makedirs(cut_folder, exist_ok=True)
        page = self.t4_page.get()
        src_path = os.path.join(folder, f"page_{page}.png")
        if not os.path.exists(src_path):
            self.t4_log_msg(f"❌ 元画像なし: {src_path}")
            messagebox.showerror("エラー", f"page_{page}.png が見つかりません。")
            return
        try:
            img = Image.open(src_path)
            dpi = 96
            top_px = mm_to_px(top_var.get() * 10, dpi)
            bottom_px = mm_to_px(bottom_var.get() * 10, dpi)
            left_px = mm_to_px(left_var.get() * 10, dpi)
            right_px = mm_to_px(right_var.get() * 10, dpi)
            w, h = img.size
            crop_left = left_px
            crop_top = top_px
            crop_right = w - right_px
            crop_bottom = h - bottom_px
            if crop_right <= crop_left or crop_bottom <= crop_top:
                self.t4_log_msg("⚠️ カット量が画像サイズを超えています")
                messagebox.showerror("エラー", "カット量が画像サイズを超えています。")
                return
            cropped = img.crop((crop_left, crop_top, crop_right, crop_bottom))
            out_name = f"page_{page}{suffix_filename}"
            out_path = os.path.join(cut_folder, out_name)
            cropped.save(out_path)
            self.t4_log_msg(f"✅ カット完了: {out_name}")
            self.status_var.set(f"カット完了: page_{page}")
            messagebox.showinfo("成功", f"カット完了！\n{out_path}")
        except Exception as e:
            self.t4_log_msg(f"❌ エラー: {e}")
            messagebox.showerror("エラー", str(e))

    def _t4_generic_resize(self, src_suffix, out_suffix, w_var, h_var):
        folder = self.t4_image_folder.get().strip()
        if not folder:
            messagebox.showerror("エラー", "画像フォルダを選択してください。")
            return
        resize_folder = os.path.join(folder, "resize")
        os.makedirs(resize_folder, exist_ok=True)
        page = self.t4_page.get()
        src_name = f"page_{page}{src_suffix}"
        src_path = os.path.join(folder, "_cut", src_name)
        if not os.path.exists(src_path):
            self.t4_log_msg(f"❌ カット画像なし: {src_path}")
            messagebox.showerror("エラー", f"先にカット実行を行ってください。\n{src_path}")
            return
        try:
            img = Image.open(src_path)
            new_w = w_var.get()
            new_h = h_var.get()
            if new_w <= 0 or new_h <= 0:
                messagebox.showerror("エラー", "幅・高さは正の整数を指定してください。")
                return
            resized = img.resize((new_w, new_h), Image.LANCZOS)
            out_name = f"page_{page}{out_suffix}"
            out_path = os.path.join(resize_folder, out_name)
            resized.save(out_path)
            self.t4_log_msg(f"✅ リサイズ完了: {out_name}")
            self.status_var.set(f"リサイズ完了: page_{page}")
            messagebox.showinfo("成功", f"リサイズ完了！\n{out_path}")
        except Exception as e:
            self.t4_log_msg(f"❌ エラー: {e}")
            messagebox.showerror("エラー", str(e))

    def t4_insert_image(self, target_col, mode):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("エラー", "openpyxlをインストールしてください。")
            return
        excel_path = self.t4_excel.get().strip()
        folder = self.t4_image_folder.get().strip()
        if not excel_path or not folder:
            messagebox.showerror("エラー", "Excelと画像フォルダを指定してください。")
            return
        page = self.t4_page.get()

        if mode == "datestamp":
            img_name = f"page_{page}_resizeDateStamp.png"
            target_cell = "D1"
        else:
            img_name = f"page_{page}_resizeDepartmentUser.png"
            target_cell = "H1"

        img_path = os.path.join(folder, "resize", img_name)
        if not os.path.exists(img_path):
            self.t4_log_msg(f"❌ リサイズ画像なし: {img_path}")
            messagebox.showerror("エラー", f"リサイズ画像が見つかりません。\n{img_path}")
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            pil_img = Image.open(img_path)

            col_letter = target_col[0]
            row_num = int(target_col[1:])
            col_w = ws.column_dimensions[col_letter].width or 8.43
            row_h = ws.row_dimensions[row_num].height or 28.5
            dpi = 96
            cell_w_px = char_to_px(col_w, dpi)
            cell_h_px = pt_to_px(row_h, dpi)
            if cell_w_px < 10:
                cell_w_px = 80
            if cell_h_px < 10:
                cell_h_px = 80

            ratio = min(cell_w_px / pil_img.width, cell_h_px / pil_img.height)
            new_w = int(pil_img.width * ratio)
            new_h = int(pil_img.height * ratio)
            if new_w < 1:
                new_w = 1
            if new_h < 1:
                new_h = 1
            resized_img = pil_img.resize((new_w, new_h), Image.LANCZOS)

            from io import BytesIO
            img_bytes = BytesIO()
            resized_img.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            xl_img = XLImage(img_bytes)
            xl_img.anchor = target_cell

            for shape in ws._images:
                if hasattr(shape, 'anchor') and hasattr(shape.anchor, '_from'):
                    if shape.anchor._from.row == row_num - 1 and shape.anchor._from.col == openpyxl.utils.column_index_from_string(col_letter) - 1:
                        ws._images.remove(shape)
            ws.add_image(xl_img)
            wb.save(excel_path)
            wb.close()
            self.t4_log_msg(f"✅ {target_cell}に挿入完了: {img_name}")
            self.status_var.set(f"挿入完了: {target_cell}")
            messagebox.showinfo("成功", f"{target_cell}に画像を挿入しました。")
        except PermissionError:
            messagebox.showerror("エラー", "Excelファイルが開かれています。閉じてから再実行してください。")
        except Exception as e:
            self.t4_log_msg(f"❌ 挿入エラー: {e}")
            messagebox.showerror("エラー", str(e))

    # ============================================================
    # タブ5: OCR（画像→テキスト化）
    # ============================================================
    def setup_tab5(self, bg_color, label_style):
        parent = self.tab5
        parent.configure(bg=bg_color)
        pad = 10

        frame = ttk.LabelFrame(parent, text="OCR（画像→テキスト化）設定", style=label_style, padding=pad)
        frame.pack(fill="x", padx=pad, pady=pad)

        row1 = tk.Frame(frame, bg=bg_color)
        row1.pack(fill="x", pady=2)
        ttk.Label(row1, text="① 画像フォルダ：").pack(side="left")
        self.ocr_source = tk.StringVar()
        ttk.Entry(row1, textvariable=self.ocr_source, width=40, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row1, text="参照", command=lambda: self.browse_folder(self.ocr_source)).pack(side="left")

        row2 = tk.Frame(frame, bg=bg_color)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="② 出力先（テキスト保存先）：").pack(side="left")
        self.ocr_output = tk.StringVar()
        ttk.Entry(row2, textvariable=self.ocr_output, width=40, font=self.base_font).pack(side="left", padx=5)
        ttk.Button(row2, text="参照", command=lambda: self.browse_folder(self.ocr_output)).pack(side="left")

        row3 = tk.Frame(frame, bg=bg_color)
        row3.pack(fill="x", pady=2)
        ttk.Label(row3, text="③ ファイルパターン（例：*.png, *.jpg）：").pack(side="left")
        self.ocr_pattern = tk.StringVar(value="*.png")
        ttk.Entry(row3, textvariable=self.ocr_pattern, width=20, font=self.base_font).pack(side="left", padx=5)
        ttk.Label(row3, text="言語（例：jpn, eng）：").pack(side="left", padx=10)
        self.ocr_lang = tk.StringVar(value="jpn")
        ttk.Entry(row3, textvariable=self.ocr_lang, width=10, font=self.base_font).pack(side="left", padx=5)

        row4 = tk.Frame(frame, bg=bg_color)
        row4.pack(fill="x", pady=2)
        ttk.Button(row4, text="▶ OCR実行（全画像をテキスト化）", command=self.run_ocr, style="Green.TButton").pack(side="left", padx=5)

        self.ocr_progress = ttk.Progressbar(parent, orient="horizontal", length=400, mode="determinate")
        self.ocr_progress.pack(fill="x", padx=pad, pady=5)

        log_frame = ttk.LabelFrame(parent, text="OCR処理ログ", padding=pad)
        log_frame.pack(fill="both", expand=True, padx=pad, pady=pad)
        self.ocr_log = tk.Text(log_frame, height=8, font=self.console_font, state="disabled", bg="#fafafa")
        self.ocr_log.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.ocr_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.ocr_log.config(yscrollcommand=scrollbar.set)

    def ocr_log_msg(self, msg):
        self.ocr_log.config(state="normal")
        self.ocr_log.insert(tk.END, msg + "\n")
        self.ocr_log.see(tk.END)
        self.ocr_log.config(state="disabled")
        self.root.update()

    def run_ocr(self):
        if not PYTESSERACT_AVAILABLE:
            messagebox.showerror("エラー", "pytesseractがインストールされていません。\npip install pytesseract")
            return
        src = self.ocr_source.get().strip()
        out = self.ocr_output.get().strip()
        pattern = self.ocr_pattern.get().strip()
        lang = self.ocr_lang.get().strip()

        if not src or not out:
            messagebox.showerror("エラー", "画像フォルダと出力先を指定してください。")
            return
        if not os.path.exists(src):
            messagebox.showerror("エラー", "画像フォルダが存在しません。")
            return
        if not os.path.exists(out):
            os.makedirs(out, exist_ok=True)

        files = []
        for f in os.listdir(src):
            if fnmatch.fnmatch(f, pattern):
                fpath = os.path.join(src, f)
                if os.path.isfile(fpath):
                    ext = os.path.splitext(f)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.webp']:
                        files.append(f)

        if not files:
            self.ocr_log_msg("❌ 対象画像がありません。")
            messagebox.showinfo("情報", "対象画像が見つかりませんでした。")
            return

        total = len(files)
        self.ocr_progress["maximum"] = total
        self.ocr_progress["value"] = 0
        self.ocr_log_msg(f"🔍 OCR開始: {total}件")

        all_text_lines = []
        md_lines = ["# OCR抽出結果", "", ""]

        for i, filename in enumerate(files, 1):
            img_path = os.path.join(src, filename)
            self.ocr_log_msg(f"⏳ 処理中: {filename} ({i}/{total})")
            try:
                img = Image.open(img_path)
                text = pytesseract.image_to_string(img, lang=lang)
                text = text.strip()
                if text:
                    all_text_lines.append(f"--- {filename} ---")
                    all_text_lines.append(text)
                    all_text_lines.append("")
                    md_lines.append(f"## {filename}")
                    md_lines.append("")
                    md_lines.append(text)
                    md_lines.append("")
                    md_lines.append("---")
                    md_lines.append("")
                    self.ocr_log_msg(f"✅ {filename}: テキスト抽出完了 ({len(text)}文字)")
                else:
                    all_text_lines.append(f"--- {filename} ---")
                    all_text_lines.append("（テキストなし）")
                    all_text_lines.append("")
                    md_lines.append(f"## {filename}")
                    md_lines.append("")
                    md_lines.append("（テキストなし）")
                    md_lines.append("---")
                    md_lines.append("")
                    self.ocr_log_msg(f"⚠️ {filename}: テキストが抽出されませんでした")
            except Exception as e:
                self.ocr_log_msg(f"❌ {filename}: エラー - {e}")
                all_text_lines.append(f"--- {filename} ---")
                all_text_lines.append(f"（エラー: {e}）")
                all_text_lines.append("")
                md_lines.append(f"## {filename}")
                md_lines.append("")
                md_lines.append(f"（エラー: {e}）")
                md_lines.append("---")
                md_lines.append("")

            self.ocr_progress["value"] = i
            self.root.update()

        txt_path = os.path.join(out, "OCR_result.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write("\n".join(all_text_lines))
        self.ocr_log_msg(f"✅ .txt保存: {txt_path}")

        md_path = os.path.join(out, "OCR_result.md")
        with open(md_path, "w", encoding="utf-8") as f:
            f.write("\n".join(md_lines))
        self.ocr_log_msg(f"✅ .md保存: {md_path}")

        self.status_var.set(f"✅ OCR完了: {total}件")
        self.ocr_progress["value"] = total
        messagebox.showinfo("成功", f"OCR完了！\n{total}件の画像を処理しました。\n\n出力:\n{txt_path}\n{md_path}")

    # ============================================================
    # 汎用ユーティリティ
    # ============================================================
    def browse_file(self, var, title, ext):
        path = filedialog.askopenfilename(title=title, filetypes=[(title, f"*{ext}")])
        if path:
            var.set(path)

    def browse_folder(self, var):
        path = filedialog.askdirectory(title="フォルダを選択")
        if path:
            var.set(path)

# ------------------------------------------------------------
# 起動
# ------------------------------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = SealLedgerDXApp(root)
    root.mainloop()
