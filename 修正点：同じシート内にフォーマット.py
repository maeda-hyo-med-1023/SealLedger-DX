    # ★★★ 修正点：同じシート内にフォーマット罫線を連続追加 ★★★
    def run_format_copy(self):
        template = self.fmt_template.get().strip()
        if not template or not os.path.exists(template):
            messagebox.showerror("エラー", "テンプレートファイルを選択してください。")
            return
        pages = self.fmt_pages.get()
        if pages < 1:
            messagebox.showerror("エラー", "ページ数は1以上を指定してください。")
            return

        dir_name = os.path.dirname(template)
        base_name = os.path.splitext(os.path.basename(template))[0]
        output = os.path.join(dir_name, f"{base_name}_generated.xlsx")
        if os.path.exists(output):
            if not messagebox.askyesno("確認", f"出力先\n{output}\nは既に存在します。上書きしますか？"):
                return

        try:
            shutil.copy2(template, output)
            wb = openpyxl.load_workbook(output)

            target_sheet = None
            for sheet_name in wb.sheetnames:
                if '校印使用台帳' in sheet_name:
                    target_sheet = wb[sheet_name]
                    break
            if target_sheet is None:
                target_sheet = wb[wb.sheetnames[0]]

            # テンプレートブロックの行範囲
            template_start_row = 1
            template_end_row = 23
            block_height = template_end_row - template_start_row + 1

            # 列範囲を動的に取得（最大列数まで）
            max_col = target_sheet.max_column

            current_max_row = target_sheet.max_row
            if current_max_row < template_end_row:
                current_max_row = template_end_row

            insert_start_row = current_max_row + 1

            from openpyxl.utils import get_column_letter

            for page_idx in range(pages):
                dest_start_row = insert_start_row + (page_idx * block_height)

                # セル値＋スタイルをコピー（全列）
                for row_offset in range(block_height):
                    src_row = template_start_row + row_offset
                    dest_row = dest_start_row + row_offset
                    for col in range(1, max_col + 1):
                        src_cell = target_sheet.cell(row=src_row, column=col)
                        dest_cell = target_sheet.cell(row=dest_row, column=col)
                        dest_cell.value = src_cell.value
                        copy_cell_style(src_cell, dest_cell)

                # 行の高さをコピー
                for row_offset in range(block_height):
                    src_row = template_start_row + row_offset
                    dest_row = dest_start_row + row_offset
                    if target_sheet.row_dimensions[src_row].height:
                        target_sheet.row_dimensions[dest_row].height = target_sheet.row_dimensions[src_row].height

                # セルの結合をコピー（全列対応）
                for merged_range in list(target_sheet.merged_cells.ranges):
                    if merged_range.min_row <= template_end_row and merged_range.max_row <= template_end_row:
                        row_offset = merged_range.min_row - template_start_row
                        new_min_row = dest_start_row + row_offset
                        new_max_row2 = dest_start_row + (merged_range.max_row - template_start_row)
                        new_min_col = merged_range.min_col
                        new_max_col = merged_range.max_col
                        target_sheet.merge_cells(
                            start_row=new_min_row,
                            start_column=new_min_col,
                            end_row=new_max_row2,
                            end_column=new_max_col
                        )

            # 印刷範囲を全列に拡張
            new_max_row = insert_start_row + (pages * block_height) - 1
            last_col_letter = get_column_letter(max_col)
            target_sheet.print_area = f"A1:{last_col_letter}{new_max_row}"

            wb.save(output)
            self.status_var.set(f"✅ 複製完了: {pages}ページを同じシート内に追加しました。")
            messagebox.showinfo("成功", f"完了！\n{output}\n同じシート内に{pages}ページ分のフォーマットを追加しました。")
        except Exception as e:
            messagebox.showerror("エラー", str(e))